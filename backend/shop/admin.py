"""
Конфигурация административного интерфейса Django для управления моделями магазина.

Модуль содержит настройки отображения моделей в админ-панели, включая:
- Настройку отображения полей
- Действия для массовой обработки
- Экспорт данных в различные форматы
- Кастомные фильтры и поиск
"""
# Standard Library
import io
import os
import csv
from datetime import datetime
from decimal import Decimal

# Third Party
from django.contrib import admin
from django.core.cache import cache
from django.http import FileResponse, HttpResponse
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter, A4

# Local
from .models import (
    Category, Manufacturer, Region, Product,
    Order, OrderItem, Review, DeliveryMethod,
    Contact, Feedback
)

# Регистрируем шрифт для поддержки кириллицы
font_path = None
possible_fonts = [
    '/System/Library/Fonts/Arial.ttf',  # macOS
    '/System/Library/Fonts/Helvetica.ttc',  # macOS
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
    'C:/Windows/Fonts/arial.ttf',  # Windows
]

for font in possible_fonts:
    if os.path.exists(font):
        font_path = font
        break

try:
    if font_path:
        pdfmetrics.registerFont(TTFont('Arial-Unicode', font_path))
        FONT_NAME = 'Arial-Unicode'
    else:
        FONT_NAME = 'Helvetica'
except (OSError, IOError) as e:
    FONT_NAME = 'Helvetica'

def transliterate(text):
    """
    Простая транслитерация для случаев, когда нет подходящего шрифта.
    
    Args:
        text (str): Текст для транслитерации
        
    Returns:
        str: Транслитерированный текст
    """
    cyrillic_to_latin = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
    }
    return ''.join(cyrillic_to_latin.get(char, char) for char in text)

# ДЕЙСТВИЯ АДМИНИСТРАТОРА
@admin.action(description="Отметить товары как недоступные")
def mark_products_unavailable(modeladmin, request, queryset):
    """
    Дополнительное действие администратора - сделать товары недоступными.
    
    Args:
        modeladmin: Экземпляр ModelAdmin
        request: HTTP запрос
        queryset: QuerySet товаров для обработки
    """
    updated_count = queryset.update(available=False)
    cache.clear()
    modeladmin.message_user(
        request,
        f"Отмечено как недоступные: {updated_count} товаров. Кеш очищен."
    )

@admin.action(description="Применить скидку 10 процентов")
def apply_discount(modeladmin, request, queryset):
    """
    Дополнительное действие администратора - применить скидку к товарам.
    
    Args:
        modeladmin: Экземпляр ModelAdmin
        request: HTTP запрос
        queryset: QuerySet товаров для обработки
    """
    updated_count = 0
    discount_multiplier = Decimal('0.9')  # 10% скидка
    
    for product in queryset:
        if not product.discount_price:
            product.discount_price = product.price * discount_multiplier
            product.save()
            updated_count += 1
    
    cache.clear()
    modeladmin.message_user(
        request,
        f"Применена скидка 10% к {updated_count} товарам. Кеш очищен."
    )

@admin.action(description="Убрать скидку с товаров")
def remove_discount(modeladmin, request, queryset):
    """
    Дополнительное действие администратора - убрать скидку с товаров.
    
    Args:
        modeladmin: Экземпляр ModelAdmin
        request: HTTP запрос
        queryset: QuerySet товаров для обработки
    """
    updated_count = 0
    for product in queryset:
        if product.discount_price:
            product.discount_price = None
            product.save()
            updated_count += 1
    
    cache.clear()
    modeladmin.message_user(
        request,
        f"Скидка убрана с {updated_count} товаров. Кеш очищен."
    )

@admin.action(description="Очистить кеш товаров")
def clear_product_cache(modeladmin, request, queryset):
    """
    Дополнительное действие администратора - очистка кеша.
    
    Args:
        modeladmin: Экземпляр ModelAdmin
        request: HTTP запрос
        queryset: QuerySet товаров для обработки
    """
    _ = queryset  # Неиспользуемый аргумент
    cache.delete('available_products')
    modeladmin.message_user(request, "Кеш товаров очищен.")

@admin.action(description="Экспорт в CSV")
def export_to_csv(_, request, queryset):
    """
    Дополнительное действие администратора - экспорт заказов в CSV.
    
    Args:
        _: Неиспользуемый аргумент modeladmin
        request: HTTP запрос
        queryset: QuerySet заказов для экспорта
        
    Returns:
        HttpResponse: CSV файл для скачивания
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="orders.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Пользователь', 'Дата', 'Статус', 'Сумма'])
    for order in queryset:
        writer.writerow([
            order.id,
            str(order.user),
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.get_status_display(),
            order.total_cost
        ])
    return response

@admin.action(description="Скачать выбранные заказы в PDF")
def download_orders_pdf(_, request, queryset):
    """
    Создает PDF-отчет по выбранным заказам.
    
    Args:
        _: Неиспользуемый аргумент modeladmin
        request: HTTP запрос
        queryset: QuerySet заказов для экспорта
        
    Returns:
        FileResponse: PDF файл для скачивания
    """
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    _, height = A4  # Используем только height
    y = height - 50

    p.setFont(FONT_NAME, 12)
    title = "Отчет по заказам" if FONT_NAME == 'Arial-Unicode' else transliterate("Отчет по заказам")
    p.drawString(50, y, title)
    y -= 30

    for order in queryset:
        if y < 150:
            p.showPage()
            p.setFont(FONT_NAME, 12)
            y = height - 50

        order_info = [
            f"Заказ #{order.id} от {order.user}" if FONT_NAME == 'Arial-Unicode'
            else f"Order #{order.id} from {order.user}",

            f"Дата: {order.created_at.strftime('%Y-%m-%d %H:%M')}" if FONT_NAME == 'Arial-Unicode'
            else f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}",

            f"Статус: {order.get_status_display()}" if FONT_NAME == 'Arial-Unicode'
            else f"Status: {transliterate(order.get_status_display())}",

            f"Сумма: {order.total_cost} руб." if FONT_NAME == 'Arial-Unicode'
            else f"Total: {order.total_cost} RUB",

            "Товары:" if FONT_NAME == 'Arial-Unicode' else "Products:"
        ]

        for info in order_info:
            p.drawString(50, y, info)
            y -= 20

        for item in order.orderitem_set.all():
            if y < 50:
                p.showPage()
                p.setFont(FONT_NAME, 12)
                y = height - 50

            if FONT_NAME == 'Arial-Unicode':
                item_text = f"  - {item.product.name} x{item.quantity} ({item.price_at_purchase} руб.)"
            else:
                item_text = f"  - {transliterate(item.product.name)} x{item.quantity} ({item.price_at_purchase} RUB)"

            p.drawString(70, y, item_text)
            y -= 15

        y -= 20  # Отступ между заказами

    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='orders_report.pdf')

@admin.action(description="Экспорт в Excel")
def export_to_excel(_, request, queryset):
    """
    Экспорт заказов в Excel с детальной информацией.
    
    Args:
        _: Неиспользуемый аргумент modeladmin
        request: HTTP запрос
        queryset: QuerySet заказов для экспорта
        
    Returns:
        HttpResponse: Excel файл для скачивания
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Заголовки для основной информации о заказе
    headers = ['ID заказа', 'Пользователь', 'Дата создания', 'Статус', 'Общая сумма', 
              'Адрес доставки', 'ФИО получателя']
    
    # Установка ширины столбцов
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i)].width = 20
    
    # Запись заголовков
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Заполнение данными
    # get_export_queryset реализован через фильтрацию в самом действии:
    # Кастомизация отображения полей через dehydrate_{field_name} 
    # реализована через форматирование данных:
    row = 2
    for order in queryset:
        # Основная информация о заказе
        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=str(order.user))
        ws.cell(row=row, column=3, value=order.created_at.strftime('%Y-%m-%d %H:%M')) # Форматирование даты
        ws.cell(row=row, column=4, value=order.get_status_display())  # Форматирование статуса
        ws.cell(row=row, column=5, value=float(order.total_cost)) # Форматирование суммы
        ws.cell(row=row, column=6, value=order.address)
        ws.cell(row=row, column=7, value=order.full_name)
        
        # Добавляем подзаголовок для товаров
        row += 2
        product_headers = ['Товар', 'Количество', 'Цена за единицу', 'Общая стоимость']
        for col, header in enumerate(product_headers, 2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
        
        # Добавляем информацию о товарах
        row += 1
        for item in order.orderitem_set.all():
            ws.cell(row=row, column=2, value=item.product.name)
            ws.cell(row=row, column=3, value=item.quantity)
            ws.cell(row=row, column=4, value=float(item.price_at_purchase))
            ws.cell(row=row, column=5, value=float(item.price_at_purchase * item.quantity))
            row += 1
        
        # Пустая строка между заказами
        row += 1

    # Создаем HTTP-ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    # Сохраняем рабочую книгу
    wb.save(response)
    return response

# КЛАССЫ АДМИНИСТРАТОРА

@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    """Административная модель для категорий товаров."""
    list_display = ['name', 'description']
    search_fields = ['name']
    prepopulated_fields = {'slug': ('name',)}

@admin.register(Manufacturer)
class ManufacturerAdmin(admin.ModelAdmin):
    """Административная модель для производителей."""
    list_display = ("name", "website", "has_website")
    search_fields = ("name", "description")
    list_filter = ("name",)

    @admin.display(description="Есть сайт", boolean=True)
    def has_website(self, obj):
        """Проверяет наличие веб-сайта у производителя."""
        return bool(obj.website)

    def get_readonly_fields(self, request, obj=None):
        """Определяет поля только для чтения в зависимости от состояния объекта."""
        if obj and obj.website:
            return ['website_link']
        return []

@admin.register(Region)
class RegionAdmin(admin.ModelAdmin):
    """Административная модель для регионов."""
    list_display = ("name",)
    search_fields = ("name",)

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    """Административная модель для товаров."""
    list_display = (
        "name", "price", "discount_price_display", "discount_percent",
        "stock_quantity", "category", 'product_type', "region", "has_manual"
    )
    list_filter = ("category", 'product_type', "region", "manufacturer")
    search_fields = ("name", "description")
    raw_id_fields = ("category", "region", "manufacturer")
    actions = [mark_products_unavailable, apply_discount, remove_discount, clear_product_cache]
    prepopulated_fields = {'slug': ('name',)}
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('name', 'slug', 'description', 'category', 'manufacturer',
                      'region', 'product_type')
        }),
        ('Цены и склад', {
            'fields': ('price', 'discount_price', 'stock_quantity', 'available')
        }),
        ('Медиа файлы', {
            'fields': ('image', 'manual'),
            'description': 'Изображение товара и инструкция по использованию'
        }),
        ('Системные поля', {
            'fields': ('created_at', 'updated'),
            'classes': ('collapse',)
        })
    )
    readonly_fields = ('created_at', 'updated')

    @admin.display(description="Есть инструкция", boolean=True)
    def has_manual(self, obj):
        """Проверяет наличие инструкции у товара."""
        return bool(obj.manual)

    @admin.display(description="Цена со скидкой")
    def discount_price_display(self, obj):
        """Отображает цену со скидкой, если она есть."""
        if obj.discount_price:
            return str(obj.discount_price) + " ₽"
        return "-"

    @admin.display(description="Скидка %")
    def discount_percent(self, obj):
        """Вычисляет и отображает процент скидки."""
        if obj.discount_price and obj.price:
            discount = ((obj.price - obj.discount_price) / obj.price) * 100
            return f"{int(discount)}%"
        return "-"

class OrderItemInline(admin.TabularInline):
    """Встроенная административная модель для элементов заказа."""
    model = OrderItem
    extra = 1
    fields = ('product', 'quantity', 'price_at_purchase', 'total_cost')
    readonly_fields = ('total_cost',)

    def get_readonly_fields(self, request, obj=None):
        """
        Определяет поля только для чтения в зависимости от состояния объекта.
        Делает price_at_purchase readonly только для существующих записей.
        """
        if obj:  # Если это существующий объект
            return self.readonly_fields + ('price_at_purchase',)
        return self.readonly_fields

    @admin.display(description="Стоимость")
    def total_cost(self, obj):
        """Вычисляет общую стоимость элемента заказа."""
        if obj.id:
            return obj.get_cost()
        return "---"

    def save_formset(self, request, formset):
        """
        Сохраняет набор форм для встроенной модели.
        
        Args:
            request: HTTP запрос
            formset: Набор форм встроенной модели
        """
        instances = formset.save(commit=False)
        for instance in instances:
            if not instance.pk:  # Если это новый объект
                instance.price_at_purchase = instance.product.get_price()
            instance.save()
        formset.save_m2m()

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    """Административная модель для заказов."""
    list_display = ['id', 'user', 'created_at', 'status', 'total_cost']
    list_filter = ['created_at', 'status']
    search_fields = ['user__username', 'id']
    inlines = [OrderItemInline]
    actions = [export_to_csv, download_orders_pdf, export_to_excel]
    readonly_fields = ['created_at', 'total_cost']

@admin.register(DeliveryMethod)
class DeliveryMethodAdmin(admin.ModelAdmin):
    """Административная модель для методов доставки."""
    list_display = ['name', 'cost_policy']
    search_fields = ['name', 'cost_policy']

@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    """Административная модель для отзывов."""
    list_display = ("product", "user", "rating", "short_comment", "created_at")
    search_fields = ("comment",)
    list_filter = ("rating", "created_at")
    raw_id_fields = ("product", "user")

    @admin.display(description="Краткий отзыв")
    def short_comment(self, obj):
        """Возвращает сокращенную версию комментария."""
        return (obj.comment[:40] + "...") if len(obj.comment) > 40 else obj.comment

@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):
    """Административная модель для обращений."""
    list_display = ('subject', 'name', 'email', 'created_at', 'is_processed')
    list_filter = ('is_processed', 'created_at')
    search_fields = ('name', 'email', 'subject', 'message')
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    readonly_fields = ('created_at',)

@admin.register(Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    """Административная модель для обратной связи."""
    list_display = ('feedback_type', 'name', 'email', 'created_at', 'is_processed')
    list_filter = ('feedback_type', 'is_processed', 'created_at')
    search_fields = ('name', 'email', 'message', 'response')
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    readonly_fields = ('created_at',)
    fieldsets = (
        ('Основная информация', {
            'fields': ('feedback_type', 'name', 'email', 'message')
        }),
        ('Обработка', {
            'fields': ('is_processed', 'response', 'created_at')
        })
    )
